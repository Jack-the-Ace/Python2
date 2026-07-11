from openpyxl import Workbook
wb= Workbook()
sheet= wb.active
sheet.title= "Scores"
#--------------------------------------------------------------------
sheet.append(['학번', '출석', '퀴즈1', '퀴즈2', '중간고사', '기말고사', '프로젝트'])
sheet.append([1, 10, 8, 5, 14, 26, 12])
sheet.append([2, 7, 3, 7, 15, 24, 18])
sheet.append([3, 9, 5, 8, 8, 12, 4])
sheet.append([4, 7, 8, 7, 17, 21, 18])
sheet.append([5, 7, 8, 7, 16, 25, 15])
sheet.append([6, 3, 5, 8, 8, 17, 0])
sheet.append([7, 4, 9, 10, 16, 27, 18])
sheet.append([8, 6, 6, 6, 15, 19, 17])
sheet.append([9, 10, 10, 9, 19, 30, 19])
sheet.append([10, 9, 8, 8, 20, 25, 20])
#--------------------------------------------------- row data 입력 끝
for d in range(2,12):
    sheet[f'D{d}'].value= 10

sheet['H1'].value='총점'
sheet['I1'].value='성적정보'

for total in range(2,12):
    sheet[f'H{total}'].value= sum(cell.value for cell in sheet[f'B{total}:G{total}'][0])

for info in range(2,12):
    if sheet[f'H{info}'].value >=90:
        sheet[f'I{info}']= 'A'
    elif sheet[f'H{info}'].value >=80:
        sheet[f'I{info}']='B'
    elif sheet[f'H{info}'].value >=70:
        sheet[f'I{info}']='C'
    else:
        sheet[f'I{info}']='D'

for absc in range(2,12):
    if int(sheet[f'B{absc}'].value)<5:
        sheet[f'I{absc}']='F'
#-------------------------------------------------------------항목추가 끝
sheet.column_dimensions['A'].width=5
sheet.row_dimensions[1].height=24

from openpyxl.styles import Font, Border, Side
for row1 in sheet['A1:I1']:
    for cell in row1:
        cell.font= Font(bold=True)

thin_side = Side(style='thin')
for row in sheet.iter_rows(min_row=1, max_row=11, min_col=1, max_col=9):
    for cell in row:
        cell.border=thin_side

thick_side=Side(style='medium')
min_row, max_row, min_col, max_col = 1, 11, 1, 9
for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
    for cell in row:       
        left = thick_side if cell.column == min_col else thin_side
        right = thick_side if cell.column == max_col else thin_side
        top = thick_side if cell.row == min_row else thin_side
        bottom = thick_side if cell.row == max_row else thin_side
        
        # 가장자리 조건에 맞는 선만 셀에 적용 (기존 테두리가 있다면 덮어쓰지 않게 조심!)
        if left or right or top or bottom:
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

#-------------------------------------------------------------스타일 끝
wb.save('./01_excel/scores.xlsx')
wb.close()